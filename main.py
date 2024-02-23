# import pandas as pd
excel_name="【考研择校】东南大学2017-2023复试分数线.xlsx"
sheet_name='2022'
from openpyxl import load_workbook
import pandas as pd

def process_merged_cells(ws):
    """
    处理合并单元格，将合并区域的第一个有效值复制到未被合并的单元格。
    避开已有的合并单元格进行赋值操作。
    """
    merged_ranges = ws.merged_cells.ranges
    all_rows, all_cols = ws.max_row, ws.max_column
    non_merged_coordinates = [(row, col) for row in range(1, all_rows + 1)
                                       for col in range(1, all_cols + 1)
                                       if (row, col) not in [cell.coordinate for cell in merged_ranges]]

    for merged_cell in merged_ranges:
        value = ws.cell(row=merged_cell.min_row, column=merged_cell.min_col).value
        
        for row, col in non_merged_coordinates:
            if merged_cell.min_row <= row <= merged_cell.max_row and merged_cell.min_col <= col <= merged_cell.max_col:
                ws.cell(row=row, column=col).value = value

# 加载Excel文件
wb = load_workbook(excel_name)
ws = wb.active  # 或者指定特定的工作表 ws = wb['Sheet1']

# 处理合并单元格
process_merged_cells(ws)

# 将工作表转换为DataFrame（此处假设所有非合并的单元格都已经填充了正确的值）
data = []
for row in ws.iter_rows(values_only=True):
    data.append(row)

df = pd.DataFrame(data)

# 转换为字典列表
data_dict_list = df.to_dict('records')

for item in data_dict_list:
    print(item)